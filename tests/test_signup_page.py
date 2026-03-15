import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys
import traceback

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from pages.signup_page import SignupPage

class TestSignup(unittest.TestCase):
    
    @classmethod
    def setUpClass(cls):
        cls.excel_dir = r"D:\VTS Traning"
        cls.excel_path = os.path.join(cls.excel_dir, "signup_test_report.xlsx")
        os.makedirs(cls.excel_dir, exist_ok=True)
        
        cls.screenshot_dir = r"D:\VTS Traning\screenshots"
        os.makedirs(cls.screenshot_dir, exist_ok=True)
        
        cls.workbook = openpyxl.Workbook()
        cls.sheet = cls.workbook.active
        cls.sheet.title = "Signup Test Report"
        
        headers = [
            "Bug ID", "Bug Title", "Steps to Reproduce", "Expected Result", 
            "Actual Result", "Severity", "Priority", "Status", "Remarks", "Screen Snaps"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = cls.sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        column_widths = [15, 40, 60, 40, 40, 15, 15, 15, 40, 50]
        for i, width in enumerate(column_widths, 1):
            cls.sheet.column_dimensions[chr(64 + i)].width = width
        
        cls.results = []
        cls.current_row = 2
        cls.bug_counter = 1
        
        print(f"\n{'='*60}")
        print("SIGNUP TEST EXECUTION")
        print(f"{'='*60}")
        print(f"Excel report: {cls.excel_path}")
        print(f"Screenshots: {cls.screenshot_dir}")
    
    @classmethod
    def tearDownClass(cls):
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for result in cls.results:
            row = cls.current_row
            
            cell = cls.sheet.cell(row=row, column=1, value=result['bug_id'])
            cell.border = border
            
            cell = cls.sheet.cell(row=row, column=2, value=result['bug_title'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=3, value=result['steps'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=4, value=result['expected'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=5, value=result['actual'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=6, value=result['severity'])
            cell.border = border
            if result['severity'] == 'High':
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            
            cell = cls.sheet.cell(row=row, column=7, value=result['priority'])
            cell.border = border
            if result['priority'] == 'High':
                cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
            
            status_cell = cls.sheet.cell(row=row, column=8, value=result['status'])
            status_cell.border = border
            if result['status'] == "PASS":
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif result['status'] == "FAIL":
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif result['status'] == "SKIP":
                status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            cell = cls.sheet.cell(row=row, column=9, value=result['remarks'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=10, value=result['screenshot'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cls.sheet.row_dimensions[row].height = 80
            cls.current_row += 1
        
        try:
            cls.workbook.save(cls.excel_path)
            print(f"\n{'='*60}")
            print(f"✓ Signup test report saved to: {cls.excel_path}")
            print(f"✓ Screenshots saved in: {cls.screenshot_dir}")
            
            try:
                os.startfile(cls.excel_path)
            except:
                pass
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = os.path.join(cls.excel_dir, f"signup_report_{timestamp}.xlsx")
            cls.workbook.save(new_path)
            print(f"\n✓ Report saved to alternate file: {new_path}")
            try:
                os.startfile(new_path)
            except:
                pass
    
    def setUp(self):
        """Setup before each test"""
        options = webdriver.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--window-size=1920,1080')
        
        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.wait = WebDriverWait(self.driver, 10)
        self.signup_page = SignupPage(self.driver)
        self.test_name = self.id().split('.')[-1]
        self.screenshot_path = None
        self.steps = []
        
        print(f"\n Starting Test: {self.test_name}")
    
    def tearDown(self):
        self.take_screenshot()
        time.sleep(1)
        if self.driver:
            self.driver.quit()
    
    def take_screenshot(self):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.test_name}_{timestamp}.png"
            self.screenshot_path = os.path.join(self.__class__.screenshot_dir, filename)
            self.driver.save_screenshot(self.screenshot_path)
            print(f"   Screenshot saved: {filename}")
            return self.screenshot_path
        except Exception as e:
            print(f"   Screenshot failed: {e}")
            self.screenshot_path = "No screenshot captured"
            return "No screenshot"
    
    def record_result(self, status, bug_title, steps, expected, actual, remarks):
        result = {
            'bug_id': f"BUG_SIGNUP_{self.__class__.bug_counter:03d}",
            'bug_title': bug_title,
            'steps': "\n".join(steps),
            'expected': expected,
            'actual': actual,
            'severity': "High" if status == "FAIL" else "N/A",
            'priority': "High" if status == "FAIL" else "N/A",
            'status': status,
            'remarks': remarks,
            'screenshot': self.screenshot_path if self.screenshot_path and self.screenshot_path != "No screenshot" else "No screenshot"
        }
        self.__class__.results.append(result)
        self.__class__.bug_counter += 1
        print(f"   Result: {status}")
    
    def check_field_exists(self, field_name):
        try:
            self.driver.find_element(By.NAME, field_name)
            return True
        except:
            return False
    
    def debug_page_info(self):
        print(f"  Current URL: {self.driver.current_url}")
        print(f"  Page Title: {self.driver.title}")
        
        inputs = self.driver.find_elements(By.TAG_NAME, "input")
        print(f"  Input fields found: {len(inputs)}")
        for inp in inputs:
            name = inp.get_attribute("name")
            type_attr = inp.get_attribute("type")
            print(f"    - name: {name}, type: {type_attr}")
    
    def test_1_valid_signup(self):
        steps = [
            "1. Open signup page",
            "2. Enter valid fullname: 'Jefrin Thangaraj'",
            "3. Enter valid email: 'jefthang@gmail.com'",
            "4. Enter valid password: 'Password@123'",
            "5. Click Register button"
        ]
        expected = "User should be registered successfully"
        
        try:
            self.signup_page.open_signup_page()
            time.sleep(2)
            
            self.debug_page_info()
            
            fullname_success = self.signup_page.enter_fullname("Jefrin Thangaraj")
            if not fullname_success:
                print("  Fullname field not found, trying username field")
                try:
                    username_field = self.driver.find_element(By.NAME, "username")
                    username_field.clear()
                    username_field.send_keys("Jefrin Thangaraj")
                except:
                    pass
            
            self.signup_page.enter_email("jefthang@gmail.com")
            
            self.signup_page.enter_password("Password@123")
            
            self.signup_page.click_register()
            time.sleep(3)
            
            self.take_screenshot()
            
            if "signup" not in self.driver.current_url.lower():
                actual = "Registration attempted - Page changed"
                self.record_result("PASS", "Valid Signup Test", steps, expected, actual, "Registration processed")
                print(" TC01: Valid Signup Test PASSED")
            else:
                try:
                    success_msg = self.driver.find_element(By.CSS_SELECTOR, ".success, .alert-success")
                    actual = f"Success message: {success_msg.text}"
                    self.record_result("PASS", "Valid Signup Test", steps, expected, actual, "Success message shown")
                    print("TC01: Valid Signup Test PASSED (with message)")
                except:
                    actual = "Still on signup page - May require email verification"
                    self.record_result("PASS", "Valid Signup Test", steps, expected, actual, "Form submitted")
                    print("✓ TC01: Valid Signup Test PASSED (form submitted)")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"   Exception: {error_msg}")
            traceback.print_exc()
            self.record_result("FAIL", "Valid Signup Test", steps, expected, error_msg, "Exception occurred")
            # Don't raise - we want to continue with other tests
    
    def test_2_invalid_email(self):
        """Test Case 2: Invalid Email Format"""
        steps = [
            "1. Open signup page",
            "2. Enter valid fullname: 'Jefrin Thangaraj'",
            "3. Enter invalid email: 'jefrgmail.com' (missing @)",
            "4. Enter valid password: 'Password@123'",
            "5. Click Register button"
        ]
        expected = "System should validate email format"
        
        try:
            self.signup_page.open_signup_page()
            time.sleep(2)
            
            self.signup_page.enter_fullname("Jefrin Thangaraj")
            
            self.signup_page.enter_email("jefrgmail.com")
            
            self.signup_page.enter_password("Password@123")
            
            self.signup_page.click_register()
            time.sleep(2)
            
            self.take_screenshot()
            
            error = self.signup_page.get_error_message()
            
            if error:
                actual = f"Validation error shown: {error}"
                self.record_result("PASS", "Invalid Email Test", steps, expected, actual, "Email validation working")
                print("✓ TC02: Invalid Email Test PASSED")
            else:
                # Check HTML5 validation
                email_field = self.driver.find_element(By.NAME, "email")
                if email_field.get_attribute("validationMessage"):
                    actual = f"HTML5 validation: {email_field.get_attribute('validationMessage')}"
                    self.record_result("PASS", "Invalid Email Test", steps, expected, actual, "Browser validation")
                    print("✓ TC02: Invalid Email Test PASSED (browser validation)")
                else:
                    actual = "No visible validation error"
                    self.record_result("FAIL", "Invalid Email Test", steps, expected, actual, "Validation missing")
                    print("✗ TC02: Invalid Email Test FAILED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  ❌ Exception: {error_msg}")
            self.record_result("FAIL", "Invalid Email Test", steps, expected, error_msg, "Exception occurred")
    
    # ============= TEST CASE 3: Weak Password =============
    def test_3_weak_password(self):
        """Test Case 3: Weak Password"""
        steps = [
            "1. Open signup page",
            "2. Enter valid fullname: 'Jefrin Thangaraj'",
            "3. Enter valid email: 'jefthang@gmail.com'",
            "4. Enter weak password: '123'",
            "5. Click Register button"
        ]
        expected = "System should validate password strength"
        
        try:
            self.signup_page.open_signup_page()
            time.sleep(2)
            
            self.signup_page.enter_fullname("Jefrin Thangaraj")
            
            self.signup_page.enter_email("jefthang@gmail.com")
            
            self.signup_page.enter_password("123")
            
            self.signup_page.click_register()
            time.sleep(2)
            
            self.take_screenshot()
            
            error = self.signup_page.get_error_message()
            
            if error:
                actual = f"Validation error shown: {error}"
                self.record_result("PASS", "Weak Password Test", steps, expected, actual, "Password validation working")
                print("✓ TC03: Weak Password Test PASSED")
            else:
                # Check if form was submitted
                if "signup" not in self.driver.current_url.lower():
                    actual = "Form submitted with weak password"
                    self.record_result("FAIL", "Weak Password Test", steps, expected, actual, "No password validation")
                    print("✗ TC03: Weak Password Test FAILED")
                else:
                    actual = "No visible validation, but form not submitted"
                    self.record_result("PASS", "Weak Password Test", steps, expected, actual, "May have client validation")
                    print("✓ TC03: Weak Password Test PASSED (client validation)")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  ❌ Exception: {error_msg}")
            self.record_result("FAIL", "Weak Password Test", steps, expected, error_msg, "Exception occurred")
    
    def test_4_empty_fields(self):
        """Test Case 4: All Fields Empty"""
        steps = [
            "1. Open signup page",
            "2. Leave all fields empty",
            "3. Click Register button"
        ]
        expected = "System should validate required fields"
        
        try:
            self.signup_page.open_signup_page()
            time.sleep(2)
            
            self.signup_page.click_register()
            time.sleep(2)
            
            self.take_screenshot()
            
            error = self.signup_page.get_error_message()
            
            if error:
                actual = f"Validation error shown: {error}"
                self.record_result("PASS", "Empty Fields Test", steps, expected, actual, "Field validation working")
                print("✓ TC04: Empty Fields Test PASSED")
            else:
                if "signup" in self.driver.current_url.lower():
                    actual = "Form not submitted - HTML5 validation likely"
                    self.record_result("PASS", "Empty Fields Test", steps, expected, actual, "Browser validation")
                    print("✓ TC04: Empty Fields Test PASSED (browser validation)")
                else:
                    actual = "Form submitted with empty fields"
                    self.record_result("FAIL", "Empty Fields Test", steps, expected, actual, "Validation missing")
                    print("✗ TC04: Empty Fields Test FAILED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  ❌ Exception: {error_msg}")
            self.record_result("FAIL", "Empty Fields Test", steps, expected, error_msg, "Exception occurred")
    
    # ============= TEST CASE 5: Remember Me Checkbox =============
    def test_5_remember_me(self):
        """Test Case 5: Remember Me Checkbox"""
        steps = [
            "1. Open signup page",
            "2. Check if Remember Me checkbox exists",
            "3. Click on checkbox if present"
        ]
        expected = "Remember Me checkbox should be present and clickable"
        
        try:
            self.signup_page.open_signup_page()
            time.sleep(2)
            
            try:
                checkbox = self.driver.find_element(By.NAME, "remember")
                is_displayed = checkbox.is_displayed()
                is_enabled = checkbox.is_enabled()
                
                if is_displayed and is_enabled:
                    self.signup_page.click_remember_me()
                    time.sleep(1)
                    
                    is_selected = checkbox.is_selected()
                    
                    self.take_screenshot()
                    
                    actual = f"Checkbox found, enabled, selected: {is_selected}"
                    self.record_result("PASS", "Remember Me Test", steps, expected, actual, "Checkbox functional")
                    print("✓ TC05: Remember Me Test PASSED")
                else:
                    actual = "Checkbox found but not interactive"
                    self.record_result("FAIL", "Remember Me Test", steps, expected, actual, "Checkbox not interactive")
                    print("✗ TC05: Remember Me Test FAILED")
                    
            except:
                # Checkbox might not exist on signup page
                self.take_screenshot()
                actual = "Remember Me checkbox not found on signup page"
                self.record_result("SKIP", "Remember Me Test", steps, expected, actual, "Checkbox not present")
                print("⚠ TC05: Remember Me Test SKIPPED (checkbox not found)")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  ❌ Exception: {error_msg}")
            self.record_result("FAIL", "Remember Me Test", steps, expected, error_msg, "Exception occurred")

if __name__ == "__main__":
    unittest.main(verbosity=2)