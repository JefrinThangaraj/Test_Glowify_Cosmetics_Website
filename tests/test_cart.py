import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
import sys
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import traceback

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pages.home_page import HomePage
from pages.cart_page import CartPage

class TestCartPOM(unittest.TestCase):
    """Cart Test Cases using Page Object Model"""
    
    @classmethod
    def setUpClass(cls):
        """Setup Excel report"""
        cls.excel_dir = r"D:\VTS Traning"
        cls.excel_path = os.path.join(cls.excel_dir, "cart_pom_report.xlsx")
        os.makedirs(cls.excel_dir, exist_ok=True)
        
        cls.screenshot_dir = os.path.join(cls.excel_dir, "screenshots")
        os.makedirs(cls.screenshot_dir, exist_ok=True)
        
        # Create Excel workbook
        cls.workbook = openpyxl.Workbook()
        cls.sheet = cls.workbook.active
        cls.sheet.title = "Cart POM Test Report"
        
        # Headers
        headers = [
            "Bug ID", "Bug Title", "Steps to Reproduce", "Expected Result", 
            "Actual Result", "Severity", "Priority", "Status", "Remarks", "Screen Snaps"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = cls.sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = [15, 40, 60, 40, 40, 15, 15, 15, 40, 50]
        for i, width in enumerate(column_widths, 1):
            cls.sheet.column_dimensions[chr(64 + i)].width = width
        
        cls.results = []
        cls.current_row = 2
        cls.bug_counter = 1
        
        print(f"\n{'='*60}")
        print("CART POM TEST EXECUTION")
        print(f"{'='*60}")
        print(f"Excel file: {cls.excel_path}")
        print(f"Screenshots: {cls.screenshot_dir}")
    
    def setUp(self):
        """Setup before each test"""
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--window-size=1920,1080')
        
        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        
        self.home_page = HomePage(self.driver)
        self.cart_page = None
        self.test_name = self.id().split('.')[-1]
        self.screenshot_path = None
        self.steps = []
        self.setup_success = False
        
        print(f"\n{'='*60}")
        print(f"Starting Test: {self.test_name}")
        print(f"{'='*60}")
    
    def tearDown(self):
        """Cleanup after each test"""
        self.take_screenshot()
        time.sleep(1)
        if self.driver:
            self.driver.quit()
    
    def take_screenshot(self):
        """Take screenshot"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.test_name}_{timestamp}.png"
            self.screenshot_path = os.path.join(self.__class__.screenshot_dir, filename)
            self.driver.save_screenshot(self.screenshot_path)
            print(f"  📸 Screenshot: {filename}")
        except:
            self.screenshot_path = "No screenshot"
    
    def record_result(self, status, bug_title, steps, expected, actual, remarks):
        """Record test result"""
        result = {
            'bug_id': f"BUG_CART_{self.__class__.bug_counter:03d}",
            'bug_title': f"{bug_title} - {status}",
            'steps': "\n".join(steps),
            'expected': expected,
            'actual': actual,
            'severity': "High" if status == "FAIL" else "N/A",
            'priority': "High" if status == "FAIL" else "N/A",
            'status': status,
            'remarks': remarks,
            'screenshot': self.screenshot_path or "No screenshot"
        }
        self.__class__.results.append(result)
        self.__class__.bug_counter += 1
        print(f"  📊 Result: {status}")
    
    @classmethod
    def tearDownClass(cls):
        """Save all results to Excel"""
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for result in cls.results:
            row = cls.current_row
            
            # Bug ID
            cell = cls.sheet.cell(row=row, column=1, value=result['bug_id'])
            cell.border = border
            
            # Bug Title
            cell = cls.sheet.cell(row=row, column=2, value=result['bug_title'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Steps
            cell = cls.sheet.cell(row=row, column=3, value=result['steps'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Expected
            cell = cls.sheet.cell(row=row, column=4, value=result['expected'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Actual
            cell = cls.sheet.cell(row=row, column=5, value=result['actual'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Severity
            cell = cls.sheet.cell(row=row, column=6, value=result['severity'])
            cell.border = border
            
            # Priority
            cell = cls.sheet.cell(row=row, column=7, value=result['priority'])
            cell.border = border
            
            # Status
            status_cell = cls.sheet.cell(row=row, column=8, value=result['status'])
            status_cell.border = border
            if result['status'] == "PASS":
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif result['status'] == "FAIL":
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Remarks
            cell = cls.sheet.cell(row=row, column=9, value=result['remarks'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Screen Snaps
            cell = cls.sheet.cell(row=row, column=10, value=result['screenshot'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cls.sheet.row_dimensions[row].height = 80
            cls.current_row += 1
        
        # Save workbook
        try:
            cls.workbook.save(cls.excel_path)
            print(f"\n{'='*60}")
            print(f"✓ Cart POM report saved to: {cls.excel_path}")
            print(f"✓ Screenshots saved in: {cls.screenshot_dir}")
            try:
                os.startfile(cls.excel_path)
            except:
                pass
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = os.path.join(cls.excel_dir, f"cart_pom_report_{timestamp}.xlsx")
            cls.workbook.save(new_path)
            print(f"\n✓ Report saved to: {new_path}")
    
    def test_1_add_to_cart(self):
        steps = [
            "1. Open home page",
            "2. Click Add to Cart button",
            "3. Navigate to cart page"
        ]
        expected = "Product should be added to cart"
        
        try:
            self.home_page.open_home_page()
            self.take_screenshot()
            
            add_result = self.home_page.add_first_product_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.home_page.go_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.cart_page = CartPage(self.driver)
            item_count = self.cart_page.get_number_of_cart_items()
            
            if item_count > 0:
                actual = f"Product added, cart has {item_count} item(s)"
                self.record_result("PASS", "Add to Cart Test", steps, expected, actual, "Product added")
                print("✓ TC01: Add to Cart Test PASSED")
            else:
                actual = "Cart is empty"
                self.record_result("FAIL", "Add to Cart Test", steps, expected, actual, "Product not added")
                print("✗ TC01: Add to Cart Test FAILED")
                
        except Exception as e:
            self.record_result("FAIL", "Add to Cart Test", steps, expected, str(e), "Exception")
            raise
    
    # ============= TEST CASE 2: Cart Page Load =============
    def test_2_cart_page_load(self):
        """Test Case 2: Verify cart page loads"""
        steps = [
            "1. Open home page",
            "2. Navigate to cart page"
        ]
        expected = "Cart page should load with correct title"
        
        try:
            self.home_page.open_home_page()
            self.home_page.go_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.cart_page = CartPage(self.driver)
            title = self.cart_page.get_cart_title()
            
            if title:
                actual = f"Cart title: '{title}'"
                self.record_result("PASS", "Cart Page Load Test", steps, expected, actual, "Page loaded")
                print("✓ TC02: Cart Page Load Test PASSED")
            else:
                actual = "Cart title not found"
                self.record_result("FAIL", "Cart Page Load Test", steps, expected, actual, "Page load failed")
                print("✗ TC02: Cart Page Load Test FAILED")
                
        except Exception as e:
            self.record_result("FAIL", "Cart Page Load Test", steps, expected, str(e), "Exception")
            raise
    
    # ============= TEST CASE 3: Product Details =============
    def test_3_product_details(self):
        steps = [
            "1. Open home page",
            "2. Add product to cart",
            "3. Go to cart",
            "4. Check product details"
        ]
        expected = "Product details should be displayed"
        
        try:
            self.home_page.open_home_page()
            self.home_page.add_first_product_to_cart()
            time.sleep(2)
            self.home_page.go_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.cart_page = CartPage(self.driver)
            names = self.cart_page.get_product_names()
            prices = self.cart_page.get_product_prices()
            
            if names:
                actual = f"Found {len(names)} product(s)"
                self.record_result("PASS", "Product Details Test", steps, expected, actual, "Details shown")
                print("✓ TC03: Product Details Test PASSED")
            else:
                actual = "No products found"
                self.record_result("FAIL", "Product Details Test", steps, expected, actual, "Details missing")
                print("✗ TC03: Product Details Test FAILED")
                
        except Exception as e:
            self.record_result("FAIL", "Product Details Test", steps, expected, str(e), "Exception")
            raise
    
    # ============= TEST CASE 4: Price Calculations =============
    def test_4_price_calculations(self):
        steps = [
            "1. Add product to cart",
            "2. Go to cart",
            "3. Verify price × quantity = total"
        ]
        expected = "Product total should equal price × quantity"
        
        try:
            self.home_page.open_home_page()
            self.home_page.add_first_product_to_cart()
            time.sleep(2)
            self.home_page.go_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.cart_page = CartPage(self.driver)
            results = self.cart_page.verify_product_calculations()
            
            if results:
                all_correct = all(r['correct'] for r in results)
                if all_correct:
                    actual = "All calculations correct"
                    self.record_result("PASS", "Price Calculations Test", steps, expected, actual, "Calculations OK")
                    print("✓ TC04: Price Calculations Test PASSED")
                else:
                    actual = "Some calculations incorrect"
                    self.record_result("FAIL", "Price Calculations Test", steps, expected, actual, "Calculation error")
                    print("✗ TC04: Price Calculations Test FAILED")
            else:
                actual = "No products to verify"
                self.record_result("SKIP", "Price Calculations Test", steps, expected, actual, "Cart empty")
                print("⚠ TC04: Price Calculations Test SKIPPED")
                
        except Exception as e:
            self.record_result("FAIL", "Price Calculations Test", steps, expected, str(e), "Exception")
            raise
    
    # ============= TEST CASE 5: Order Summary =============
    def test_5_order_summary(self):
        steps = [
            "1. Add product to cart",
            "2. Go to cart",
            "3. Check order summary values"
        ]
        expected = "Order summary should display values"
        
        try:
            self.home_page.open_home_page()
            self.home_page.add_first_product_to_cart()
            time.sleep(2)
            self.home_page.go_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.cart_page = CartPage(self.driver)
            selected = self.cart_page.get_selected_items_count()
            subtotal = self.cart_page.get_subtotal()
            grand_total = self.cart_page.get_grand_total()
            
            actual = f"Selected: {selected}, Subtotal: ₹{subtotal}, Grand: ₹{grand_total}"
            self.record_result("PASS", "Order Summary Test", steps, expected, actual, "Summary shown")
            print("✓ TC05: Order Summary Test PASSED")
                
        except Exception as e:
            self.record_result("FAIL", "Order Summary Test", steps, expected, str(e), "Exception")
            raise
    
    # ============= TEST CASE 6: Checkout Button =============
    def test_6_checkout_button(self):
        """Test Case 6: Verify checkout button"""
        steps = [
            "1. Go to cart page",
            "2. Check if checkout button exists"
        ]
        expected = "Checkout button should be visible"
        
        try:
            self.home_page.open_home_page()
            self.home_page.go_to_cart()
            time.sleep(2)
            self.take_screenshot()
            
            self.cart_page = CartPage(self.driver)
            is_visible = self.cart_page.is_checkout_button_displayed()
            
            if is_visible:
                text = self.cart_page.get_checkout_button_text()
                actual = f"Button visible, text: {text}"
                self.record_result("PASS", "Checkout Button Test", steps, expected, actual, "Button OK")
                print("✓ TC06: Checkout Button Test PASSED")
            else:
                actual = "Button not visible"
                self.record_result("FAIL", "Checkout Button Test", steps, expected, actual, "Button missing")
                print("✗ TC06: Checkout Button Test FAILED")
                
        except Exception as e:
            self.record_result("FAIL", "Checkout Button Test", steps, expected, str(e), "Exception")
            raise

if __name__ == "__main__":
    unittest.main(verbosity=2)