import unittest
from selenium import webdriver
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Import all test classes
from tests.test_login import TestLogin
from tests.test_signup_page import TestSignupPage
from tests.test_cart import TestCartSummary

class MasterTestRunner:
    """Run all tests and store results in single Excel file"""
    
    def __init__(self):
        self.excel_dir = r"D:\VTS Traning"
        self.excel_path = os.path.join(self.excel_dir, "master_test_report.xlsx")
        self.screenshot_dir = os.path.join(self.excel_dir, "screenshots")
        
        # Create directories
        os.makedirs(self.excel_dir, exist_ok=True)
        os.makedirs(self.screenshot_dir, exist_ok=True)
        
        # Setup Excel
        self.setup_excel()
        
        # Store all results
        self.all_results = []
        self.current_row = 2
    
    def setup_excel(self):
        """Setup Excel workbook with headers"""
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Master Test Report"
        
        headers = [
            "Bug ID", "Bug Title", "Steps to Reproduce", "Expected Result", 
            "Actual Result", "Severity", "Priority", "Status", "Remarks", "Screen Snaps"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = [15, 40, 60, 40, 40, 15, 15, 15, 40, 50]
        for i, width in enumerate(column_widths, 1):
            self.sheet.column_dimensions[chr(64 + i)].width = width
    
    def save_results_to_excel(self):
        """Save all results to Excel"""
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for result in self.all_results:
            row = self.current_row
            
            # Bug ID
            cell = self.sheet.cell(row=row, column=1, value=result.get('bug_id', 'N/A'))
            cell.border = border
            
            # Bug Title
            cell = self.sheet.cell(row=row, column=2, value=result.get('bug_title', 'N/A'))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Steps to Reproduce
            cell = self.sheet.cell(row=row, column=3, value=result.get('steps', 'N/A'))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Expected Result
            cell = self.sheet.cell(row=row, column=4, value=result.get('expected', 'N/A'))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Actual Result
            cell = self.sheet.cell(row=row, column=5, value=result.get('actual', 'N/A'))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Severity
            severity_cell = self.sheet.cell(row=row, column=6, value=result.get('severity', 'N/A'))
            severity_cell.border = border
            if result.get('severity') == 'High':
                severity_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            
            # Priority
            priority_cell = self.sheet.cell(row=row, column=7, value=result.get('priority', 'N/A'))
            priority_cell.border = border
            if result.get('priority') == 'High':
                priority_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
            
            # Status
            status_cell = self.sheet.cell(row=row, column=8, value=result.get('status', 'N/A'))
            status_cell.border = border
            if result.get('status') == "PASS":
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif result.get('status') == "FAIL":
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Remarks
            cell = self.sheet.cell(row=row, column=9, value=result.get('remarks', 'N/A'))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Screen Snaps
            cell = self.sheet.cell(row=row, column=10, value=result.get('screenshot', 'N/A'))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            self.sheet.row_dimensions[row].height = 80
            self.current_row += 1
        
        # Save workbook
        try:
            self.workbook.save(self.excel_path)
            print(f"\n✓ Master test report saved to: {self.excel_path}")
            os.startfile(self.excel_path)
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = os.path.join(self.excel_dir, f"master_report_{timestamp}.xlsx")
            self.workbook.save(new_path)
            print(f"\n✓ Report saved to: {new_path}")
            os.startfile(new_path)
    
    def run(self):
        """Run all tests"""
        print("=" * 70)
        print("MASTER TEST EXECUTION - GLOWIFY COSMETICS")
        print("=" * 70)
        
        # Set common screenshot directory for all test classes
        TestLogin.screenshot_dir = self.screenshot_dir
        TestSignupPage.screenshot_dir = self.screenshot_dir
        TestCartSummary.screenshot_dir = self.screenshot_dir
        
        # Create test suite
        loader = unittest.TestLoader()
        suite = unittest.TestSuite()
        
        # Add all tests
        suite.addTests(loader.loadTestsFromTestCase(TestLogin))
        suite.addTests(loader.loadTestsFromTestCase(TestSignupPage))
        suite.addTests(loader.loadTestsFromTestCase(TestCartSummary))
        
        # Run tests
        runner = unittest.TextTestRunner(verbosity=2)
        result = runner.run(suite)
        
        # Collect results from test classes
        if hasattr(TestLogin, 'results'):
            self.all_results.extend(TestLogin.results)
        if hasattr(TestSignupPage, 'results'):
            self.all_results.extend(TestSignupPage.results)
        if hasattr(TestCartSummary, 'results'):
            self.all_results.extend(TestCartSummary.results)
        
        # Save to Excel
        self.save_results_to_excel()
        
        return result

if __name__ == "__main__":
    runner = MasterTestRunner()
    runner.run()