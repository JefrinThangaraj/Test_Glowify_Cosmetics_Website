import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys
import traceback

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from pages.login_page import LoginPage

class TestLogin(unittest.TestCase):
    
    @classmethod
    def setUpClass(cls):
        cls.excel_dir = r"D:\VTS Traning"
        cls.excel_path = os.path.join(cls.excel_dir, "login_test_report.xlsx")
        os.makedirs(cls.excel_dir, exist_ok=True)
        
        cls.screenshot_dir = r"D:\VTS Traning\screenshots"
        os.makedirs(cls.screenshot_dir, exist_ok=True)
        
        cls.workbook = openpyxl.Workbook()
        cls.sheet = cls.workbook.active
        cls.sheet.title = "Login Test Report"
        
        headers = [
            "Bug ID", "Bug Title", "Steps to Reproduce", "Expected Result", 
            "Actual Result", "Severity", "Priority", "Status", "Remarks", "Screen Snaps"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = cls.sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        column_widths = [15, 40, 60, 40, 40, 15, 15, 15, 40, 50]
        for i, width in enumerate(column_widths, 1):
            cls.sheet.column_dimensions[chr(64 + i)].width = width
        
        cls.results = []
        cls.current_row = 2
        cls.bug_counter = 1
        
        print(f"\n{'='*60}")
        print("LOGIN TEST EXECUTION")
        print(f"{'='*60}")
        print(f"Excel report: {cls.excel_path}")
        print(f"Screenshots: {cls.screenshot_dir}")
    
    @classmethod
    def tearDownClass(cls):
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for result in cls.results:
            row = cls.current_row
            
            cell = cls.sheet.cell(row=row, column=1, value=result['bug_id'])
            cell.border = border
            
            cell = cls.sheet.cell(row=row, column=2, value=result['bug_title'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=3, value=result['steps'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=4, value=result['expected'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=5, value=result['actual'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=6, value=result['severity'])
            cell.border = border
            
            cell = cls.sheet.cell(row=row, column=7, value=result['priority'])
            cell.border = border
            
            status_cell = cls.sheet.cell(row=row, column=8, value=result['status'])
            status_cell.border = border
            if result['status'] == "PASS":
                status_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif result['status'] == "FAIL":
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif result['status'] == "SKIP":
                status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            cell = cls.sheet.cell(row=row, column=9, value=result['remarks'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cell = cls.sheet.cell(row=row, column=10, value=result['screenshot'])
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            cls.sheet.row_dimensions[row].height = 80
            cls.current_row += 1
        
        try:
            cls.workbook.save(cls.excel_path)
            print(f"\n{'='*60}")
            print(f" Login test report saved to: {cls.excel_path}")
            print(f" Screenshots saved in: {cls.screenshot_dir}")
            try:
                os.startfile(cls.excel_path)
            except:
                pass
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = os.path.join(cls.excel_dir, f"login_report_{timestamp}.xlsx")
            cls.workbook.save(new_path)
            print(f"\n Report saved to: {new_path}")
    
    def setUp(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--window-size=1920,1080')
        
        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.login_page = LoginPage(self.driver)
        self.screenshot_path = None
        self.test_name = self.id().split('.')[-1]
        self.steps = []
        
        print(f"\n Starting Test: {self.test_name}")
    
    def tearDown(self):
        self.take_screenshot()
        time.sleep(1)
        if self.driver:
            self.driver.quit()
    
    def take_screenshot(self):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.test_name}_{timestamp}.png"
            self.screenshot_path = os.path.join(self.__class__.screenshot_dir, filename)
            self.driver.save_screenshot(self.screenshot_path)
            print(f" Screenshot saved: {filename}")
            return self.screenshot_path
        except Exception as e:
            print(f"  Screenshot failed: {e}")
            self.screenshot_path = "No screenshot captured"
            return "No screenshot"
    
    def record_result(self, status, bug_title, steps, expected, actual, remarks):
        """Record test result with screenshot path"""
        result = {
            'bug_id': f"BUG_LOGIN_{self.__class__.bug_counter:03d}",
            'bug_title': bug_title,
            'steps': "\n".join(steps),
            'expected': expected,
            'actual': actual,
            'severity': "High" if status == "FAIL" else "N/A",
            'priority': "High" if status == "FAIL" else "N/A",
            'status': status,
            'remarks': remarks,
            'screenshot': self.screenshot_path if self.screenshot_path and self.screenshot_path != "No screenshot captured" else "No screenshot"
        }
        self.__class__.results.append(result)
        self.__class__.bug_counter += 1
        print(f" Result: {status}")
    
    def test_1_valid_login(self):
        steps = [
            "1. Open login page",
            "2. Enter valid email: 'jefthang@gmail.com'",
            "3. Enter valid password: 'Password@123'",
            "4. Click Login button"
        ]
        expected = "User should be logged in successfully"
        
        try:
            self.login_page.open_url("https://glowify-cosmetics-site.onrender.com/accounts/login/")
            time.sleep(2)
            
            self.take_screenshot()
            
            self.login_page.enter_user_email("jefthang@gmail.com")
            self.login_page.enter_user_password("Password@123")
            self.login_page.click_login()
            time.sleep(3)
            
            self.take_screenshot()
            
            if "login" not in self.driver.current_url.lower():
                actual = "Login successful - Redirected from login page"
                self.record_result("PASS", "Valid Login Test", steps, expected, actual, "User logged in")
                print(" TC01: Valid Login Test PASSED")
            else:
                actual = "Still on login page - Check credentials"
                self.record_result("FAIL", "Valid Login Test", steps, expected, actual, "Login failed")
                print(" TC01: Valid Login Test FAILED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  Exception: {error_msg}")
            traceback.print_exc()
            self.record_result("FAIL", "Valid Login Test", steps, expected, error_msg, "Exception occurred")
    
    def test_2_invalid_email(self):
        steps = [
            "1. Open login page",
            "2. Enter invalid email: 'jefthanggmail.com' (missing @)",
            "3. Enter password: 'Password@123'",
            "4. Click Login button"
        ]
        expected = "System should show validation error for invalid email"
        
        try:
            self.login_page.open_url("https://glowify-cosmetics-site.onrender.com/accounts/login/")
            time.sleep(2)
            
            self.login_page.enter_user_email("jefthanggmail.com")
            self.login_page.enter_user_password("Password@123")
            self.login_page.click_login()
            time.sleep(2)
            
            self.take_screenshot()
            
            try:
                error = self.driver.find_element(By.CSS_SELECTOR, ".errorlist, .alert-danger")
                actual = f"Error shown: {error.text}"
                self.record_result("PASS", "Invalid Email Test", steps, expected, actual, "Validation working")
                print("TC02: Invalid Email Test PASSED")
            except:
                email_field = self.driver.find_element(By.NAME, "email")
                if email_field.get_attribute("validationMessage"):
                    actual = f"HTML5 validation: {email_field.get_attribute('validationMessage')}"
                    self.record_result("PASS", "Invalid Email Test", steps, expected, actual, "Browser validation")
                    print(" TC02: Invalid Email Test PASSED (browser)")
                else:
                    actual = "No error message displayed"
                    self.record_result("FAIL", "Invalid Email Test", steps, expected, actual, "Validation missing")
                    print(" TC02: Invalid Email Test FAILED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f" Exception: {error_msg}")
            self.record_result("FAIL", "Invalid Email Test", steps, expected, error_msg, "Exception occurred")
    
    def test_3_wrong_password(self):
        steps = [
            "1. Open login page",
            "2. Enter email: 'jefthang@gmail.com'",
            "3. Enter wrong password: 'WrongPass123'",
            "4. Click Login button"
        ]
        expected = "System should show error for wrong password"
        
        try:
            self.login_page.open_url("https://glowify-cosmetics-site.onrender.com/accounts/login/")
            time.sleep(2)
            
            self.login_page.enter_user_email("jefthang@gmail.com")
            self.login_page.enter_user_password("WrongPass123")
            self.login_page.click_login()
            time.sleep(2)
            
            self.take_screenshot()
            
            try:
                error = self.driver.find_element(By.CSS_SELECTOR, ".errorlist, .alert-danger")
                actual = f"Error shown: {error.text}"
                self.record_result("PASS", "Wrong Password Test", steps, expected, actual, "Validation working")
                print("TC03: Wrong Password Test PASSED")
            except:
                actual = "No error message displayed"
                self.record_result("FAIL", "Wrong Password Test", steps, expected, actual, "Validation missing")
                print("TC03: Wrong Password Test FAILED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  Exception: {error_msg}")
            self.record_result("FAIL", "Wrong Password Test", steps, expected, error_msg, "Exception occurred")
    
    def test_4_empty_fields(self):
        steps = [
            "1. Open login page",
            "2. Leave both fields empty",
            "3. Click Login button"
        ]
        expected = "System should validate required fields"
        
        try:
            self.login_page.open_url("https://glowify-cosmetics-site.onrender.com/accounts/login/")
            time.sleep(2)
            
            self.take_screenshot()
            
            self.login_page.click_login()
            time.sleep(2)
            
            self.take_screenshot()
            
            if "login" in self.driver.current_url.lower():
                actual = "Form not submitted - Validation present"
                self.record_result("PASS", "Empty Fields Test", steps, expected, actual, "Validation working")
                print("TC04: Empty Fields Test PASSED")
            else:
                actual = "Form submitted with empty fields"
                self.record_result("FAIL", "Empty Fields Test", steps, expected, actual, "Validation missing")
                print("TC04: Empty Fields Test FAILED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f" Exception: {error_msg}")
            self.record_result("FAIL", "Empty Fields Test", steps, expected, error_msg, "Exception occurred")
    
    def test_5_remember_me(self):
        steps = [
            "1. Open login page",
            "2. Enter email: 'jefthang@gmail.com'",
            "3. Enter password: 'Password@123'",
            "4. Click Remember Me checkbox",
            "5. Click Login"
        ]
        expected = "Remember Me checkbox should be clickable"
        
        try:
            self.login_page.open_url("https://glowify-cosmetics-site.onrender.com/accounts/login/")
            time.sleep(2)
            
            self.login_page.enter_user_email("jefthang@gmail.com")
            self.login_page.enter_user_password("Password@123")
            
            self.take_screenshot()
            
            try:
                checkbox = self.driver.find_element(By.NAME, "remember")
                is_displayed = checkbox.is_displayed()
                is_enabled = checkbox.is_enabled()
                
                if is_displayed and is_enabled:
                    self.login_page.click_rementer()
                    time.sleep(1)
                    
                    is_selected = checkbox.is_selected()
                    
                    self.take_screenshot()
                    
                    actual = f"Checkbox found, enabled, selected: {is_selected}"
                    self.record_result("PASS", "Remember Me Test", steps, expected, actual, "Checkbox functional")
                    print("TC05: Remember Me Test PASSED")
                else:
                    actual = "Checkbox found but not interactive"
                    self.record_result("FAIL", "Remember Me Test", steps, expected, actual, "Checkbox issue")
                    print("TC05: Remember Me Test FAILED")
            except:
                self.take_screenshot()
                actual = "Remember Me checkbox not found"
                self.record_result("SKIP", "Remember Me Test", steps, expected, actual, "Checkbox not present")
                print("TC05: Remember Me Test SKIPPED")
                
        except Exception as e:
            self.take_screenshot()
            error_msg = str(e)
            print(f"  Exception: {error_msg}")
            self.record_result("FAIL", "Remember Me Test", steps, expected, error_msg, "Exception occurred")

if __name__ == "__main__":
    unittest.main(verbosity=2)